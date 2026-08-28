"""Asynchronous, cache-first logistics sales workflow."""

from __future__ import annotations

import asyncio
from copy import copy
from io import BytesIO
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import text

from app.config import settings
from app.db.session import SessionLocal
from app.workflows.logistics_sales_workflow import (
    LogisticsSalesWorkflow,
    LogisticsWorkflowError,
    ProductKey,
    WorkflowPlan,
)
from app.workflows.sales_cache import DailySalesRecord
from app.workflows.postgres_sales_cache import PostgresSalesCache


class CachedLogisticsSalesWorkflow(LogisticsSalesWorkflow):
    """Run long LingXing syncs in a background task and compute from PostgreSQL."""

    CACHE_BATCH_SIZE = 40
    QUERY_GAP_SECONDS = 1.1

    def __init__(self):
        super().__init__()
        self.cache = PostgresSalesCache()
        self._tasks: dict[str, asyncio.Task[Any]] = {}
        self._job_results: dict[str, dict[str, Any]] = {}
        self._execute_results: dict[str, dict[str, Any]] = {}
        self._export_results: dict[str, dict[str, Any]] = {}
        self._export_files: dict[str, tuple[bytes, str, datetime]] = {}
        # A single process-wide gate protects the MCP QPS=1 contract, including
        # overlapping preview requests started by different browser sessions.
        self._sync_lock = asyncio.Lock()

    async def start_preview(self, force_refresh: bool = False) -> dict[str, Any]:
        message = "等待从领星强制刷新" if force_refresh else "等待后台同步"
        job_id = self.cache.create_job(message)
        task = asyncio.create_task(self._run_preview(job_id, force_refresh))
        self._tasks[job_id] = task
        return {
            "status": "queued",
            "job_id": job_id,
            "progress": 0,
            "message": (
                "任务已创建，准备从领星强制刷新"
                if force_refresh
                else "任务已创建，等待后台同步"
            ),
            "error": "",
            "preview": None,
        }

    async def job_status(self, job_id: str) -> dict[str, Any]:
        self._purge_exports(datetime.now(ZoneInfo("Asia/Shanghai")))
        job = self.cache.job(job_id)
        if job is None:
            raise LogisticsWorkflowError("同步任务不存在或已过期")
        result = self._job_results.get(job_id)
        return {
            "status": job["status"],
            "job_id": job_id,
            "progress": job["progress"],
            "message": job["message"],
            "error": job["error"],
            "preview": result,
            "result": self._execute_results.get(job_id),
            "export": self._export_results.get(job_id),
        }

    async def start_export(self, force_refresh: bool = False) -> dict[str, Any]:
        """Queue a temporary workbook export sourced entirely from PostgreSQL."""
        message = "等待从领星强制刷新后生成 Excel" if force_refresh else "等待读取 PostgreSQL 数据"
        job_id = self.cache.create_job(message)
        task = asyncio.create_task(self._run_export(job_id, force_refresh))
        self._tasks[job_id] = task
        return {
            "status": "queued",
            "job_id": job_id,
            "progress": 0,
            "message": message,
            "error": "",
            "preview": None,
            "result": None,
            "export": None,
        }

    async def _run_export(self, job_id: str, force_refresh: bool = False) -> None:
        async with self._sync_lock:
            await self._run_export_locked(job_id, force_refresh)

    async def _run_export_locked(self, job_id: str, force_refresh: bool = False) -> None:
        """Refresh missing sales data, then build an in-memory downloadable xlsx."""
        try:
            self.cache.update_job(job_id, "running", 2, "正在读取 PostgreSQL 商品数据")
            groups, product_rows, read_warnings, total_rows = await asyncio.to_thread(
                self._read_database_products
            )
            if not groups:
                raise LogisticsWorkflowError("PostgreSQL 中没有可导出的商品数据")

            unique_skus = sorted({key.sku for key in groups if key.sku})
            today = datetime.now(ZoneInfo("Asia/Shanghai")).date()
            # Sync the full number of completed days requested (today excluded).
            sync_start = today - timedelta(days=settings.LOGISTICS_INITIAL_SYNC_DAYS)
            sync_end = today - timedelta(days=1)
            await self._ensure_cache(
                job_id,
                unique_skus,
                sync_start,
                sync_end,
                force_refresh=force_refresh,
            )
            self.cache.update_job(job_id, "running", 92, "正在计算排除当天的滚动销量")
            updates, _, match_warnings, missing_rows = await asyncio.to_thread(
                self._build_cached_updates,
                groups,
                today,
            )
            warnings = [self._platform_scope_warning()]
            warnings.extend(read_warnings)
            warnings.extend(match_warnings)
            duplicate_groups = sum(1 for rows in groups.values() if len(rows) > 1)
            self.cache.update_job(job_id, "running", 97, "正在生成临时 Excel")
            workbook_bytes = await asyncio.to_thread(
                self._build_export_workbook,
                product_rows,
                groups,
                updates,
                warnings,
                today,
            )
            expires_at = datetime.now(ZoneInfo("Asia/Shanghai")) + timedelta(
                minutes=self.PREVIEW_TTL_MINUTES
            )
            filename = f"备货逻辑看板表-{today:%Y%m%d}.xlsx"
            self._export_files[job_id] = (workbook_bytes, filename, expires_at)
            result = {
                "status": "success",
                "filename": filename,
                "download_url": f"/api/workflows/logistics/sales-to-stock-sheet/export/{job_id}/download",
                "sheet": self.SHEET_NAME,
                "target_columns": "AJ:AM",
                "total_rows": total_rows,
                "matched_rows": len(updates),
                "missing_rows": missing_rows,
                "duplicate_groups": duplicate_groups,
                "warnings": warnings,
                "expires_at": expires_at.isoformat(timespec="seconds"),
            }
            self._export_results[job_id] = result
            self.cache.update_job(job_id, "complete", 100, "Excel 已生成，可下载")
        except Exception as exc:
            self.cache.update_job(job_id, "failed", 100, "Excel 生成失败", str(exc))

    def get_export_file(self, job_id: str) -> tuple[bytes, str] | None:
        """Return a still-valid artifact without writing it to persistent storage."""
        now = datetime.now(ZoneInfo("Asia/Shanghai"))
        self._purge_exports(now)
        artifact = self._export_files.get(job_id)
        if artifact is None:
            return None
        content, filename, _ = artifact
        return content, filename

    def _purge_exports(self, now: datetime) -> None:
        for job_id, (_, _, expires_at) in list(self._export_files.items()):
            if expires_at <= now:
                self._export_files.pop(job_id, None)
                self._export_results.pop(job_id, None)

    def _export_template_path(self) -> Path:
        configured = settings.LOGISTICS_EXPORT_TEMPLATE_PATH
        path = Path(configured) if configured else Path(__file__).resolve().parents[2] / "templates" / "备货逻辑看板表.xlsx"
        if not path.exists() or not path.is_file():
            raise LogisticsWorkflowError(f"找不到完整备货表模板：{path}")
        return path

    def _read_database_products(
        self,
    ) -> tuple[dict[ProductKey, list[int]], dict[int, dict[str, Any]], list[dict[str, Any]], int]:
        """Read product master data and the latest replenishment inputs.

        ``source_row`` is used when it points at a row in the reference
        workbook.  Missing/out-of-range/colliding source rows are assigned the
        next available row, and the mapping is reported to the user.
        """
        template = load_workbook(self._export_template_path(), read_only=True, data_only=False)
        try:
            template_max_row = template["Sheet1"].max_row if "Sheet1" in template.sheetnames else 374
        finally:
            template.close()

        query = text(
            """
            SELECT
                p.sku, p.amazon_sku, p.product_name, p.category, p.store,
                m.name AS country, p.source_row,
                p.unit_length_cm, p.unit_width_cm, p.unit_height_cm,
                p.unit_weight_g, p.unit_weight_kg,
                p.carton_length_cm, p.carton_width_cm, p.carton_height_cm,
                p.carton_quantity, p.carton_weight_kg, p.carton_volume_m3,
                p.unit_name, p.purchase_cost_cny, p.production_lead_days,
                p.supplier,
                setting.listing_date, setting.lifecycle_status,
                setting.forecast_daily_sales_override,
                setting.warehouse_split_ratio, setting.notes,
                policy.channel_type, policy.west_coast_ocean_days,
                policy.east_coast_ocean_days, policy.listing_days,
                policy.fbm_to_fba_transfer_days, policy.overall_buffer_days,
                policy.inventory_warning_ratio,
                inventory.snapshot_at, inventory.on_hand_qty,
                inventory.in_transit_qty, inventory.supplier_reserved_qty,
                inventory.west_warehouse_qty, inventory.east_warehouse_qty
            FROM product_market_parameters AS p
            JOIN market_codes AS m ON m.code = p.country_code
            LEFT JOIN replenishment_item_settings AS setting
              ON setting.product_market_parameter_id = p.id
             AND setting.is_active
            LEFT JOIN LATERAL (
                SELECT selected_policy.*
                FROM replenishment_policies AS selected_policy
                WHERE selected_policy.is_active
                  AND selected_policy.effective_from <= CURRENT_DATE
                  AND (selected_policy.effective_to IS NULL OR selected_policy.effective_to >= CURRENT_DATE)
                  AND (selected_policy.id = setting.policy_id OR selected_policy.policy_code = 'DEFAULT')
                ORDER BY (selected_policy.id = setting.policy_id) DESC
                LIMIT 1
            ) AS policy ON TRUE
            LEFT JOIN LATERAL (
                SELECT snapshot.*
                FROM inventory_position_snapshots AS snapshot
                WHERE snapshot.product_market_parameter_id = p.id
                ORDER BY snapshot.snapshot_at DESC
                LIMIT 1
            ) AS inventory ON TRUE
            WHERE p.is_active
            ORDER BY p.source_row NULLS LAST, p.sku, m.name, p.id
            """
        )
        with SessionLocal() as session:
            rows = session.execute(query).mappings().all()

        groups: dict[ProductKey, list[int]] = {}
        product_rows: dict[int, dict[str, Any]] = {}
        warnings: list[dict[str, Any]] = []
        used_rows: set[int] = set()
        next_row = 2

        def allocate_row(source_row: Any) -> tuple[int, bool]:
            nonlocal next_row
            try:
                candidate = int(source_row) if source_row is not None else 0
            except (TypeError, ValueError):
                candidate = 0
            if 2 <= candidate <= template_max_row and candidate not in used_rows:
                used_rows.add(candidate)
                return candidate, False
            while next_row in used_rows:
                next_row += 1
            assigned = next_row
            used_rows.add(assigned)
            next_row += 1
            return assigned, True

        for row in rows:
            row_number, remapped = allocate_row(row["source_row"])
            data = dict(row)
            data["_export_row"] = row_number
            product_rows[row_number] = data
            key = ProductKey.from_values(
                (row["sku"], row["amazon_sku"], row["product_name"], row["category"], row["store"], row["country"])
            )
            if remapped:
                warnings.append(
                    {
                        "level": "warning",
                        "code": "source_row_remapped",
                        "message": "数据库 source_row 无法直接对应模板行，已自动分配导出行",
                        "rows": [row_number],
                        "identity": key.as_dict(),
                    }
                )
            if not key.is_complete():
                warnings.append(
                    {
                        "level": "warning",
                        "code": "incomplete_identity",
                        "message": "数据库商品的 A-F 标识不完整，该行不会写入销量",
                        "rows": [row_number],
                        "identity": key.as_dict(),
                    }
                )
            groups.setdefault(key, []).append(row_number)

        for key, rows_for_key in groups.items():
            if len(rows_for_key) > 1:
                warnings.append(
                    {
                        "level": "warning",
                        "code": "duplicate_identity",
                        "message": "A-F 完全相同，将向这些行写入相同销量",
                        "rows": rows_for_key,
                        "identity": key.as_dict(),
                    }
                )
        return groups, product_rows, warnings, len(rows)

    def _build_export_workbook(
        self,
        product_rows: dict[int, dict[str, Any]],
        groups: dict[ProductKey, list[int]],
        updates: dict[int, tuple[int, int, int, int]],
        warnings: list[dict[str, Any]],
        today: date,
    ) -> bytes:
        """Fill the bundled full-format template with DB values and formulas."""
        workbook = load_workbook(self._export_template_path(), keep_links=False, data_only=False)
        if "Sheet1" not in workbook.sheetnames:
            raise LogisticsWorkflowError("完整备货表模板缺少 Sheet1 工作表")
        sheet = workbook["Sheet1"]
        template_max_row = sheet.max_row
        original: dict[int, dict[int, Any]] = {
            row: {column: sheet.cell(row, column).value for column in range(1, 56)}
            for row in range(2, template_max_row + 1)
        }
        max_row = max(template_max_row, max(product_rows, default=1))

        def as_number(value: Any) -> Any:
            if value is None:
                return None
            try:
                number = float(value)
            except (TypeError, ValueError):
                return value
            return int(number) if number.is_integer() else number

        def put(row: int, column: int, value: Any) -> None:
            sheet.cell(row, column).value = as_number(value)

        # Clear stale rows from the copied template, but leave all styles,
        # widths, hidden columns and the other explanation sheets intact.
        for row in range(2, max_row + 1):
            if row > template_max_row and template_max_row >= 2:
                for column in range(1, 56):
                    source = sheet.cell(template_max_row, column)
                    target = sheet.cell(row, column)
                    if source.has_style:
                        target._style = copy(source._style)
                    if source.number_format:
                        target.number_format = source.number_format
                    if source.alignment:
                        target.alignment = copy(source.alignment)
            for column in range(1, 56):
                sheet.cell(row, column).value = None

        # Columns whose original template formula encodes a per-product policy.
        # They are retained where possible; external-link formulas are replaced
        # with safe local formulas/values below.
        preserve_formula_columns = (24, 25, 26, 30, 31, 32, 33, 34, 35, 46, 47, 48)
        for row, data in product_rows.items():
            old = original.get(row, {})
            unit_weight_kg = data.get("unit_weight_kg")
            if unit_weight_kg is None and data.get("unit_weight_g") is not None:
                unit_weight_kg = float(data["unit_weight_g"]) / 1000
            # A-F identity and master/product attributes from PostgreSQL.
            for column, value in enumerate(
                (data.get("sku"), data.get("amazon_sku"), data.get("product_name"), data.get("category"), data.get("store"), data.get("country")),
                start=1,
            ):
                put(row, column, value)
            for column, value in {
                7: data.get("unit_length_cm"), 8: data.get("unit_width_cm"), 9: data.get("unit_height_cm"),
                11: unit_weight_kg, 12: data.get("carton_length_cm"), 13: data.get("carton_width_cm"),
                14: data.get("carton_height_cm"), 15: data.get("carton_quantity"), 16: data.get("carton_weight_kg"),
                17: data.get("carton_volume_m3"), 18: data.get("unit_name"), 19: data.get("listing_date"),
                20: data.get("lifecycle_status"), 21: data.get("supplier"), 22: data.get("purchase_cost_cny"),
                23: data.get("production_lead_days"), 27: data.get("on_hand_qty", 0), 28: data.get("in_transit_qty", 0),
                29: data.get("supplier_reserved_qty", 0), 45: data.get("forecast_daily_sales_override"),
                49: data.get("notes"), 51: data.get("west_warehouse_qty", 0), 52: data.get("east_warehouse_qty", 0),
                55: data.get("warehouse_split_ratio"),
            }.items():
                put(row, column, value)

            # Keep dimensions and downstream calculations live in Excel.
            sheet.cell(row, 10).value = f'=IF(K{row}="","",K{row}*1000)'
            if all(data.get(field) is not None for field in ("carton_length_cm", "carton_width_cm", "carton_height_cm")):
                sheet.cell(row, 17).value = f'=IFERROR(ROUND((L{row}*M{row}*N{row})/1000000,3),"")'
            else:
                put(row, 17, data.get("carton_volume_m3"))
            sales = updates.get(row)
            if sales is not None:
                for column, value in zip(range(36, 40), sales):
                    put(row, column, value)
                sheet.cell(row, 40).value = f"=ROUND(AJ{row}/3,0)"
                sheet.cell(row, 41).value = f"=ROUND(AK{row}/7,0)"
                sheet.cell(row, 42).value = f"=ROUND(AL{row}/15,0)"
                sheet.cell(row, 43).value = f"=ROUND(AM{row}/30,0)"
                sheet.cell(row, 44).value = (
                    f'=ROUND(IF(AND(AJ{row}>0,AK{row}>0,AL{row}>0,AM{row}>0),'
                    f'(AJ{row}/3+AK{row}/7+AL{row}/15+AM{row}/30)/4,'
                    f'LOOKUP(2,1/(AN{row}:AQ{row}<>""),AN{row}:AQ{row})),0)'
                )
            else:
                # No matching LingXing result: leave AJ:AM blank and avoid
                # propagating #DIV/0! through the calculated columns.
                for column in range(36, 45):
                    sheet.cell(row, column).value = None

            # AS is an optional manual forecast override in PostgreSQL.  When
            # it is absent, expose the calculated AR daily sales as the
            # forecast rather than leaving a misleading blank column.
            if data.get("forecast_daily_sales_override") is None:
                sheet.cell(row, 45).value = f'=IF(AR{row}="","",AR{row})'

            # Restore non-external template formulas (for example product-
            # specific safety-stock offsets), and provide DB-backed fallbacks
            # for rows that were blank in the template.
            west_days = int(data.get("west_coast_ocean_days") or 35)
            listing_days = int(data.get("listing_days") or 10)
            fbm_transfer_days = int(data.get("fbm_to_fba_transfer_days") or 45)
            buffer_days = int(data.get("overall_buffer_days") or 60)
            warning_ratio = as_number(data.get("inventory_warning_ratio") or 1.5)
            for column in preserve_formula_columns:
                value = old.get(column)
                if isinstance(value, str) and value.startswith("=") and "[" not in value:
                    sheet.cell(row, column).value = value
            fallback = {
                24: f'=IFERROR(AR{row}*(W{row}+{west_days}+{listing_days}),"")',
                25: f'=IFERROR(AR{row}*(W{row}+{west_days}+{listing_days}+{fbm_transfer_days}),"")',
                26: f'=AA{row}+AB{row}+AC{row}',
                30: f'=AA{row}+AB{row}',
                31: f'=IF(AD{row}<=AT{row},"库存不足",IF(AD{row}>=AT{row}*{warning_ratio},"库存预警","库存正常"))',
                32: f'=MAX(AT{row}-AA{row}-AB{row},0)',
                33: f'=MAX(X{row}-AD{row},0)',
                34: f'=AF{row}-AG{row}',
                35: f'=IFERROR(AX{row}+(AA{row}+AB{row})/AR{row}-{buffer_days},"")',
                46: f'=IFERROR(AR{row}*(W{row}+{listing_days}+7+{west_days}),"")',
                47: f'=IFERROR(AX{row}+AD{row}/AR{row},"")',
                48: f'=MAX(0,AT{row}-AD{row}-AC{row})',
            }
            for column, formula in fallback.items():
                if sheet.cell(row, column).value in (None, ""):
                    sheet.cell(row, column).value = formula
            # Use a guarded formula when there is no matched sales result.
            if sales is None:
                for column in (24, 25, 31, 32, 33, 34, 35, 46, 47, 48):
                    formula = sheet.cell(row, column).value
                    if isinstance(formula, str) and formula.startswith("="):
                        sheet.cell(row, column).value = f'=IFERROR({formula[1:]},"")'
            sheet.cell(row, 50).value = "=TODAY()"

            # Split recommended shipment by the configured warehouse ratio.
            sheet.cell(row, 53).value = f'=IF(OR(AF{row}="",BC{row}=""),"",IFERROR(MAX(0,AF{row}*BC{row}),""))'
            sheet.cell(row, 54).value = f'=IF(OR(AF{row}="",BC{row}=""),"",IFERROR(MAX(0,AF{row}-BA{row}),""))'

        header_fill = PatternFill("solid", fgColor="16324F")
        header_font = Font(color="FFFFFF", bold=True)
        notes = workbook.create_sheet("导出说明") if "导出说明" not in workbook.sheetnames else workbook["导出说明"]
        notes.delete_rows(1, notes.max_row)
        notes.sheet_view.showGridLines = False
        notes.append(["项目", "内容"])
        notes.append(["数据来源", "PostgreSQL（临时生成，不写入服务器文件）"])
        notes.append(["统计截止", f"{today.isoformat()}（不包含当天）"])
        notes.append(["生成时间", datetime.now(ZoneInfo("Asia/Shanghai")).isoformat(timespec="seconds")])
        notes.append(["计算说明", "AJ:AM 为排除当天的滚动销量；AR 按参考表公式计算；AA:AC、AY:AZ 来自最新库存快照；BA/BB 按 BC 分仓比例计算（BC 为空时留空）"])
        notes.append(["提示数量", len(warnings)])
        notes.append([])
        notes.append(["级别", "代码", "说明", "行号", "领星SKU"])
        for warning in warnings:
            identity = warning.get("identity") or {}
            notes.append([warning.get("level", "warning"), warning.get("code", "other"), warning.get("message", ""), "、".join(str(row) for row in warning.get("rows", [])), identity.get("sku") or identity.get("lingxing_sku", "")])
        for cell in notes[1]:
            cell.fill = header_fill
            cell.font = header_font
        notes.column_dimensions["A"].width = 14
        notes.column_dimensions["B"].width = 28
        notes.column_dimensions["C"].width = 70
        notes.column_dimensions["D"].width = 16
        notes.column_dimensions["E"].width = 18
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    async def start_execute(self, preview_id: str) -> dict[str, Any]:
        now = datetime.now(ZoneInfo("Asia/Shanghai"))
        self._purge_expired_previews(now)
        if preview_id not in self._previews:
            raise LogisticsWorkflowError("预览不存在或已过期，请重新生成预览")
        job_id = self.cache.create_job("等待写入备货表")
        task = asyncio.create_task(self._run_execute(job_id, preview_id))
        self._tasks[job_id] = task
        return {
            "status": "queued",
            "job_id": job_id,
            "progress": 0,
            "message": "写入任务已创建，等待执行",
            "error": "",
            "preview": None,
            "result": None,
        }

    async def _run_execute(self, job_id: str, preview_id: str) -> None:
        try:
            self.cache.update_job(job_id, "running", 5, "正在校验预览文件")
            now = datetime.now(ZoneInfo("Asia/Shanghai"))
            self._purge_expired_previews(now)
            plan = self._previews.get(preview_id)
            if plan is None:
                raise LogisticsWorkflowError("预览不存在或已过期，请重新生成预览")
            if self._file_hash(plan.workbook_path) != plan.workbook_hash:
                self._previews.pop(preview_id, None)
                raise LogisticsWorkflowError("预览后测试表已发生变化，为避免错行写入请重新预览")
            self.cache.update_job(job_id, "running", 10, "正在写入并校验备货表")
            async with self._write_lock:
                await asyncio.to_thread(self._write_workbook, plan)
            result = {
                "status": "success",
                "workbook": plan.workbook_path.name,
                "sheet": self.SHEET_NAME,
                "target_columns": "AJ:AM",
                "updated_rows": len(plan.updates),
                "skipped_rows": plan.missing_rows,
                "duplicate_groups": plan.duplicate_groups,
                "warnings": plan.warnings,
                "updated_at": datetime.now(ZoneInfo("Asia/Shanghai")).isoformat(timespec="seconds"),
            }
            self._execute_results[job_id] = result
            self._previews.pop(preview_id, None)
            self.cache.update_job(job_id, "complete", 100, "备货表写入完成")
        except Exception as exc:
            self.cache.update_job(job_id, "failed", 100, "写入失败", str(exc))

    async def _run_preview(self, job_id: str, force_refresh: bool = False) -> None:
        async with self._sync_lock:
            await self._run_preview_locked(job_id, force_refresh)

    async def _run_preview_locked(self, job_id: str, force_refresh: bool = False) -> None:
        try:
            self.cache.update_job(job_id, "running", 2, "正在读取测试表")
            target = self._target_path()
            groups, read_warnings, total_rows = await asyncio.to_thread(
                self._read_product_groups,
                target,
            )
            unique_skus = sorted({key.sku for key in groups})
            today = datetime.now(ZoneInfo("Asia/Shanghai")).date()
            sync_start = today - timedelta(days=settings.LOGISTICS_INITIAL_SYNC_DAYS)
            sync_end = today - timedelta(days=1)
            await self._ensure_cache(
                job_id,
                unique_skus,
                sync_start,
                sync_end,
                force_refresh=force_refresh,
            )
            self.cache.update_job(job_id, "running", 92, "正在按排除当天的规则计算销量")
            updates, keys_by_row, match_warnings, missing_rows = await asyncio.to_thread(
                self._build_cached_updates,
                groups,
                today,
            )
            warnings = [self._platform_scope_warning()]
            warnings.extend(read_warnings)
            warnings.extend(match_warnings)
            duplicate_groups = sum(1 for rows in groups.values() if len(rows) > 1)
            now = datetime.now(ZoneInfo("Asia/Shanghai"))
            preview_id = f"preview-{job_id}"
            plan = WorkflowPlan(
                preview_id=preview_id,
                workbook_path=target,
                workbook_hash=self._file_hash(target),
                expires_at=now + timedelta(minutes=self.PREVIEW_TTL_MINUTES),
                updates=updates,
                keys_by_row=keys_by_row,
                total_rows=total_rows,
                unique_products=len(groups),
                missing_rows=missing_rows,
                duplicate_groups=duplicate_groups,
                warnings=warnings,
            )
            self._previews[preview_id] = plan
            result = {
                "status": "preview",
                "preview_id": preview_id,
                "workbook": target.name,
                "sheet": self.SHEET_NAME,
                "target_columns": "AJ:AM",
                "total_rows": total_rows,
                "unique_products": len(groups),
                "matched_rows": len(updates),
                "missing_rows": missing_rows,
                "duplicate_groups": duplicate_groups,
                "warnings": warnings,
                "can_execute": bool(updates),
                "expires_at": plan.expires_at.isoformat(timespec="seconds"),
            }
            self._job_results[job_id] = result
            self.cache.update_job(job_id, "complete", 100, "预览已生成")
        except Exception as exc:
            self.cache.update_job(job_id, "failed", 100, "同步失败", str(exc))

    async def _ensure_cache(
        self,
        job_id: str,
        skus: list[str],
        start: date,
        end: date,
        force_refresh: bool = False,
    ) -> None:
        if force_refresh:
            wanted_dates = self.cache.date_range(start, end)
            missing = {sku: wanted_dates for sku in sorted(set(skus))}
        else:
            missing = self.cache.missing_dates(skus, start, end)
            # Yesterday can still be corrected in LingXing. Refresh it before
            # calculating rolling periods even when coverage is marked complete.
            for sku in sorted(set(skus)):
                dates = missing.setdefault(sku, [])
                if end not in dates:
                    dates.append(end)
        ranges: dict[tuple[date, date], list[str]] = defaultdict(list)
        for sku, dates in missing.items():
            for range_start, range_end in self._contiguous_ranges(dates):
                ranges[(range_start, range_end)].append(sku)
        requests = [
            (range_start, range_end, sku_batch)
            for (range_start, range_end), range_skus in sorted(ranges.items())
            for sku_batch in self._chunks(sorted(range_skus), self.CACHE_BATCH_SIZE)
        ]
        if not requests:
            self.cache.update_job(job_id, "running", 90, "PostgreSQL 销量缓存已覆盖所需日期")
            return
        for index, (range_start, range_end, sku_batch) in enumerate(requests, start=1):
            progress = 5 + int(index / len(requests) * 84)
            self.cache.update_job(
                job_id,
                "running",
                progress,
                f"正在查询第 {index}/{len(requests)} 批，SKU {len(sku_batch)} 个",
            )
            if index > 1:
                await asyncio.sleep(self.QUERY_GAP_SECONDS)
            request = {
                "offset": 0,
                "length": self.QUERY_PAGE_SIZE,
                "start_date": range_start.isoformat(),
                "end_date": range_end.isoformat(),
                "date_type": "purchase",
                "search_field": "local_sku",
                "search_value": sku_batch,
                "summary_field": "sku",
                "summary_field_level1": "sku",
                "turn_on_summary": 1,
                "date_view_order_type": 1,
                "date_view_type": "day",
                "sort_field": "volume",
                "sort_type": "desc",
            }
            payload = await self.client.call_tool(
                "query_product_performance_asin_lists",
                request,
            )
            records, trace_id = self._extract_daily_records(
                payload,
                range_start,
                range_end,
            )
            self.cache.save_response(request, payload, trace_id)
            self.cache.save_daily_records(
                records,
                sku_batch,
                range_start,
                range_end,
                trace_id,
                # The queried range is authoritative, including zero-sales
                # responses, so remove stale rows from that exact range first.
                replace_existing=True,
            )

    @staticmethod
    def _contiguous_ranges(dates: list[date]) -> list[tuple[date, date]]:
        if not dates:
            return []
        ordered = sorted(dates)
        ranges: list[tuple[date, date]] = []
        start = previous = ordered[0]
        for current in ordered[1:]:
            if current != previous + timedelta(days=1):
                ranges.append((start, previous))
                start = current
            previous = current
        ranges.append((start, previous))
        return ranges

    def _extract_daily_records(
        self,
        payload: Any,
        start: date,
        end: date,
    ) -> tuple[list[DailySalesRecord], str]:
        data = self._validated_payload_data(payload)
        trace_id = self._find_trace_id(payload)
        if not isinstance(data, dict):
            return [], trace_id
        source_rows = data.get("list")
        if not isinstance(source_rows, list):
            return [], trace_id
        records: list[DailySalesRecord] = []
        for row in source_rows:
            if not isinstance(row, dict):
                continue
            children = row.get("price_list")
            candidates = children if isinstance(children, list) and children else [row]
            for child in candidates:
                if not isinstance(child, dict):
                    continue
                merged = {**row, **child}
                sales_date = self._first_text(
                    merged,
                    ("rdate", "date", "sales_date", "report_date", "stat_date"),
                )
                if not sales_date and start == end:
                    sales_date = start.isoformat()
                if not sales_date:
                    continue
                sales_date = sales_date[:10]
                volume = self._first_number(merged, self.VOLUME_ALIASES)
                sku = self._first_text(merged, self.FIELD_ALIASES["sku"])
                if not sku or volume is None:
                    continue
                if volume < 0:
                    raise LogisticsWorkflowError("领星返回了负数销量，已中止")
                records.append(
                    DailySalesRecord(
                        sales_date=sales_date,
                        sku=sku,
                        amazon_sku=self._first_text(
                            merged, self.FIELD_ALIASES["amazon_sku"]
                        ),
                        product_name=self._first_text(
                            merged, self.FIELD_ALIASES["product_name"]
                        ),
                        category=self._first_text(merged, self.FIELD_ALIASES["category"]),
                        store=self._first_text(merged, self.FIELD_ALIASES["store"]),
                        country=self._first_text(merged, self.FIELD_ALIASES["country"]),
                        platform=self._first_text(
                            merged, ("platform", "platform_name", "platformName")
                        ),
                        volume=volume,
                        trace_id=trace_id,
                    )
                )
        return records, trace_id

    @staticmethod
    def _find_trace_id(payload: Any) -> str:
        if isinstance(payload, dict):
            for key in ("trace_id", "traceId", "request_id", "requestId"):
                value = payload.get(key)
                if value:
                    return str(value)
            for value in payload.values():
                trace_id = CachedLogisticsSalesWorkflow._find_trace_id(value)
                if trace_id:
                    return trace_id
        elif isinstance(payload, list):
            for value in payload:
                trace_id = CachedLogisticsSalesWorkflow._find_trace_id(value)
                if trace_id:
                    return trace_id
        return ""

    def _build_cached_updates(
        self,
        groups: dict[ProductKey, list[int]],
        today: date,
    ) -> tuple[
        dict[int, tuple[int, int, int, int]],
        dict[int, ProductKey],
        list[dict[str, Any]],
        int,
    ]:
        updates: dict[int, tuple[int, int, int, int]] = {}
        keys_by_row: dict[int, ProductKey] = {}
        warnings: list[dict[str, Any]] = []
        missing_rows = 0
        for key, rows in groups.items():
            records = self.cache.daily_records(
                key.sku,
                today - timedelta(days=settings.LOGISTICS_INITIAL_SYNC_DAYS),
                today - timedelta(days=1),
            )
            candidates = [record for record in records if record.country == key.country]
            selected = self._select_candidate(key, candidates)
            if selected is None:
                missing_rows += len(rows)
                warnings.append(
                    {
                        "level": "warning",
                        "code": "product_not_found",
                        "message": "缓存或领星结果中没有找到可唯一匹配的商品，保留原值",
                        "rows": rows,
                        "identity": {**key.as_dict(), "lingxing_sku": key.sku},
                    }
                )
                continue
            if selected[1]:
                warnings.append(
                    {
                        "level": "warning",
                        "code": "dimension_validation",
                        "message": "已按领星SKU+国家确定商品，但品名/品类/店铺存在映射差异，请复核",
                        "rows": rows,
                        "identity": {**key.as_dict(), "lingxing_sku": key.sku},
                    }
                )
            candidate = selected[0]
            sales: list[int] = []
            for days in self.PERIODS:
                start = today - timedelta(days=days)
                sales.append(
                    sum(
                        record.volume
                        for record in records
                        if record.country == key.country
                        and self._same_candidate(record, candidate)
                        and start.isoformat() <= record.sales_date < today.isoformat()
                    )
                )
            values = tuple(sales)
            for row_number in rows:
                updates[row_number] = values
                keys_by_row[row_number] = key
        return updates, keys_by_row, warnings, missing_rows

    def _select_candidate(
        self,
        key: ProductKey,
        records: list[DailySalesRecord],
    ) -> tuple[DailySalesRecord, bool] | None:
        if not records:
            return None
        candidates: dict[tuple[str, str, str, str, str], DailySalesRecord] = {}
        for record in records:
            identity = (
                record.amazon_sku,
                record.product_name,
                record.category,
                record.store,
                record.platform,
            )
            candidates.setdefault(identity, record)
        scored: list[tuple[int, DailySalesRecord, bool]] = []
        for candidate in candidates.values():
            score = 0
            mismatch = False
            if key.amazon_sku and candidate.amazon_sku == key.amazon_sku:
                score += 100
            for expected, actual in (
                (key.product_name, candidate.product_name),
                (key.category, candidate.category),
                (key.store, candidate.store),
            ):
                if expected and actual:
                    if self._normalize_text(expected) == self._normalize_text(actual):
                        score += 10
                    else:
                        mismatch = True
            scored.append((score, candidate, mismatch))
        scored.sort(key=lambda item: item[0], reverse=True)
        if len(scored) > 1 and scored[0][0] == scored[1][0]:
            return None
        return scored[0][1], scored[0][2]

    @staticmethod
    def _same_candidate(left: DailySalesRecord, right: DailySalesRecord) -> bool:
        return (
            left.amazon_sku,
            left.product_name,
            left.category,
            left.store,
            left.platform,
        ) == (
            right.amazon_sku,
            right.product_name,
            right.category,
            right.store,
            right.platform,
        )

    @staticmethod
    def _normalize_text(value: str) -> str:
        return "".join(str(value).lower().split()).replace("（", "(").replace("）", ")")


cached_logistics_sales_workflow = CachedLogisticsSalesWorkflow()
