"""Preview and write rolling LingXing sales into a local stock workbook copy."""

from __future__ import annotations

import asyncio
import errno
import hashlib
import logging
import os
import posixpath
import re
import shutil
import tempfile
import uuid
import zipfile
from xml.etree import ElementTree
from dataclasses import dataclass
from datetime import datetime, timedelta
from numbers import Number
from pathlib import Path
from typing import Any, Iterable
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from app.config import settings
from app.core.mcp_client import McpError, StreamableHttpMcpClient


class LogisticsWorkflowError(RuntimeError):
    pass


logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class ProductKey:
    sku: str
    amazon_sku: str
    product_name: str
    category: str
    store: str
    country: str

    @classmethod
    def from_values(cls, values: Iterable[Any]) -> "ProductKey":
        normalized = [str(value or "").strip() for value in values]
        if len(normalized) != 6:
            raise ValueError("A-F 商品标识必须包含 6 个字段")
        return cls(*normalized)

    def is_complete(self) -> bool:
        return all(
            (
                self.sku,
                self.product_name,
                self.category,
                self.store,
                self.country,
            )
        )

    def as_dict(self) -> dict[str, str]:
        return {
            "sku": self.sku,
            "amazon_sku": self.amazon_sku,
            "product_name": self.product_name,
            "category": self.category,
            "store": self.store,
            "country": self.country,
        }


@dataclass
class WorkflowPlan:
    preview_id: str
    workbook_path: Path
    workbook_hash: str
    expires_at: datetime
    updates: dict[int, tuple[int, int, int, int]]
    keys_by_row: dict[int, ProductKey]
    total_rows: int
    unique_products: int
    missing_rows: int
    duplicate_groups: int
    warnings: list[dict[str, Any]]


class LogisticsSalesWorkflow:
    PERIODS = (3, 7, 15, 30)
    SALES_COLUMNS = ("AJ", "AK", "AL", "AM")
    SHEET_NAME = "Sheet1"
    EXPECTED_HEADERS = (
        "领星SKU",
        "亚马逊SKU",
        "品名",
        "品类",
        "店铺",
        "国家",
    )
    PREVIEW_TTL_MINUTES = 15
    SKU_BATCH_SIZE = 40
    QUERY_PAGE_SIZE = 1000
    QUERY_FIELD = "msku"
    _write_lock = asyncio.Lock()

    FIELD_ALIASES = {
        "sku": ("local_sku", "localSku", "sku"),
        "amazon_sku": (
            "seller_sku",
            "sellerSku",
            "msku",
            "merchant_sku",
            "merchantSku",
        ),
        "product_name": (
            "local_name",
            "localName",
            "product_name",
            "productName",
            "item_name",
            "itemName",
        ),
        "category": ("category_name", "categoryName", "category"),
        "store": (
            "seller_name",
            "sellerName",
            "shop_name",
            "shopName",
            "store_name",
            "storeName",
        ),
        "country": (
            "country_name",
            "countryName",
            "country",
            "marketplace_name",
            "marketplaceName",
            "site_name",
            "siteName",
        ),
    }
    VOLUME_ALIASES = ("volume", "sales_volume", "salesVolume", "quantity")

    def __init__(self):
        self.client = StreamableHttpMcpClient(
            settings.LINGXING_MCP_URL,
            settings.LINGXING_MCP_KEY,
        )
        self._previews: dict[str, WorkflowPlan] = {}

    async def preview(self) -> dict[str, Any]:
        target = self._target_path()
        groups, read_warnings, total_rows = await asyncio.to_thread(
            self._read_product_groups,
            target,
        )
        period_sales, query_warnings = await self._fetch_period_sales(groups)
        updates, keys_by_row, match_warnings, missing_rows = self._build_updates(
            groups,
            period_sales,
        )

        warnings = [self._platform_scope_warning()]
        warnings.extend(read_warnings)
        warnings.extend(query_warnings)
        warnings.extend(match_warnings)
        duplicate_groups = sum(1 for rows in groups.values() if len(rows) > 1)

        now = datetime.now(ZoneInfo("Asia/Shanghai"))
        preview_id = uuid.uuid4().hex
        plan = WorkflowPlan(
            preview_id=preview_id,
            workbook_path=target,
            workbook_hash=self._file_hash(target),
            expires_at=now + timedelta(minutes=self.PREVIEW_TTL_MINUTES),
            updates=updates,
            keys_by_row=keys_by_row,
            total_rows=total_rows,
            unique_products=len(groups),
            missing_rows=missing_rows,
            duplicate_groups=duplicate_groups,
            warnings=warnings,
        )
        self._purge_expired_previews(now)
        self._previews[preview_id] = plan

        return {
            "preview_id": preview_id,
            "workbook": target.name,
            "sheet": self.SHEET_NAME,
            "target_columns": "AJ:AM",
            "total_rows": total_rows,
            "unique_products": len(groups),
            "matched_rows": len(updates),
            "missing_rows": missing_rows,
            "duplicate_groups": duplicate_groups,
            "warnings": warnings,
            "can_execute": bool(updates),
            "expires_at": plan.expires_at.isoformat(timespec="seconds"),
        }

    async def execute(self, preview_id: str) -> dict[str, Any]:
        now = datetime.now(ZoneInfo("Asia/Shanghai"))
        self._purge_expired_previews(now)
        plan = self._previews.get(preview_id)
        if plan is None:
            raise LogisticsWorkflowError("预览不存在或已过期，请重新生成预览")
        if self._file_hash(plan.workbook_path) != plan.workbook_hash:
            self._previews.pop(preview_id, None)
            raise LogisticsWorkflowError("预览后测试表已发生变化，为避免错行写入请重新预览")

        async with self._write_lock:
            await asyncio.to_thread(self._write_workbook, plan)
        self._previews.pop(preview_id, None)

        return {
            "workbook": plan.workbook_path.name,
            "sheet": self.SHEET_NAME,
            "target_columns": "AJ:AM",
            "updated_rows": len(plan.updates),
            "skipped_rows": plan.missing_rows,
            "duplicate_groups": plan.duplicate_groups,
            "warnings": plan.warnings,
            "updated_at": datetime.now(ZoneInfo("Asia/Shanghai")).isoformat(
                timespec="seconds"
            ),
        }

    def _target_path(self) -> Path:
        target_text = settings.STOCK_WORKBOOK_PATH.strip()
        if not target_text:
            raise LogisticsWorkflowError("未配置 STOCK_WORKBOOK_PATH")
        if target_text.startswith("\\\\") and not settings.NETWORK_WORKBOOK_WRITE_ENABLED:
            raise LogisticsWorkflowError("网络备货表写入已关闭，请配置本地测试表路径")
        target = Path(target_text)
        if not target.exists() or not target.is_file():
            raise LogisticsWorkflowError(f"找不到本地测试表：{target_text}")
        return target

    def _read_product_groups(
        self,
        target: Path,
    ) -> tuple[dict[ProductKey, list[int]], list[dict[str, Any]], int]:
        try:
            workbook = load_workbook(target, read_only=True, data_only=False)
        except Exception as exc:
            raise LogisticsWorkflowError("本地测试表不是可读取的 xlsx 文件") from exc
        try:
            if self.SHEET_NAME not in workbook.sheetnames:
                raise LogisticsWorkflowError(f"工作表 {self.SHEET_NAME} 不存在")
            sheet = workbook[self.SHEET_NAME]
            rows = sheet.iter_rows(min_row=1, max_col=6, values_only=True)
            headers = tuple(str(value or "").strip() for value in next(rows))
            if headers != self.EXPECTED_HEADERS:
                raise LogisticsWorkflowError(
                    "A-F 表头不符合预期，已中止：" + " / ".join(headers)
                )

            groups: dict[ProductKey, list[int]] = {}
            warnings: list[dict[str, Any]] = []
            total_rows = 0
            for row_number, values in enumerate(rows, start=2):
                if not any(value not in (None, "") for value in values):
                    continue
                total_rows += 1
                key = ProductKey.from_values(values)
                if not key.is_complete():
                    warnings.append(
                        {
                            "level": "warning",
                            "code": "incomplete_identity",
                            "message": "A-F 商品标识不完整，该行不会写入",
                            "rows": [row_number],
                            "identity": key.as_dict(),
                        }
                    )
                    continue
                groups.setdefault(key, []).append(row_number)

            for key, rows in groups.items():
                if len(rows) > 1:
                    warnings.append(
                        {
                            "level": "warning",
                            "code": "duplicate_identity",
                            "message": "A-F 完全相同，将向这些行写入相同销量",
                            "rows": rows,
                            "identity": key.as_dict(),
                        }
                    )
            return groups, warnings, total_rows
        finally:
            workbook.close()

    async def _fetch_period_sales(
        self,
        groups: dict[ProductKey, list[int]],
    ) -> tuple[dict[int, dict[ProductKey, int]], list[dict[str, Any]]]:
        today = datetime.now(ZoneInfo("Asia/Shanghai")).date()
        unique_mskus = sorted({key.amazon_sku for key in groups if key.amazon_sku})
        period_sales: dict[int, dict[ProductKey, int]] = {}
        warnings: list[dict[str, Any]] = []
        call_index = 0

        for days in self.PERIODS:
            sales_for_period: dict[ProductKey, int] = {}
            incomplete_records = 0
            duplicate_records = 0
            start = today - timedelta(days=days - 1)
            for msku_batch in self._chunks(unique_mskus, self.SKU_BATCH_SIZE):
                if call_index:
                    # LingXing documents a per-tool QPS limit of 1.
                    await asyncio.sleep(1.05)
                call_index += 1
                payload = await self.client.call_tool(
                    "query_product_performance_asin_lists",
                    {
                        "offset": 0,
                        "length": self.QUERY_PAGE_SIZE,
                        "start_date": start.isoformat(),
                        "end_date": today.isoformat(),
                        "date_type": "purchase",
                        "search_field": self.QUERY_FIELD,
                        "search_value": msku_batch,
                        # LingXing requires summary mode for MSKU-dimension queries.
                        "summary_field": self.QUERY_FIELD,
                        "summary_field_level1": self.QUERY_FIELD,
                        "turn_on_summary": 1,
                        "date_view_order_type": 0,
                        "sort_field": "volume",
                        "sort_type": "desc",
                    },
                )
                extracted, incomplete, duplicate = self._extract_sales_map(payload)
                incomplete_records += incomplete
                duplicate_records += duplicate
                for key, volume in extracted.items():
                    sales_for_period[key] = sales_for_period.get(key, 0) + volume

            period_sales[days] = sales_for_period
            if incomplete_records:
                warnings.append(
                    {
                        "level": "warning",
                        "code": "lingxing_incomplete_identity",
                        "message": (
                            f"近 {days} 天数据中有 {incomplete_records} 条记录缺少 A-F 对应字段，"
                            "这些记录未参与匹配"
                        ),
                        "rows": [],
                    }
                )
            if duplicate_records:
                warnings.append(
                    {
                        "level": "warning",
                        "code": "lingxing_aggregated_records",
                        "message": (
                            f"近 {days} 天数据中有 {duplicate_records} 条 A-F 重复记录，"
                            "销量已按相同商品合计"
                        ),
                        "rows": [],
                    }
                )
        return period_sales, warnings

    def _extract_sales_map(
        self,
        payload: Any,
    ) -> tuple[dict[ProductKey, int], int, int]:
        data = self._validated_payload_data(payload)
        result: dict[ProductKey, int] = {}
        incomplete_records = 0
        duplicate_records = 0

        for record in self._walk_dicts(data):
            volume = self._first_number(record, self.VOLUME_ALIASES)
            sku = self._first_text(record, self.FIELD_ALIASES["sku"])
            if volume is None or not sku:
                continue
            if volume < 0:
                raise LogisticsWorkflowError("领星返回了负数销量，已中止")
            key = ProductKey(
                sku=sku,
                amazon_sku=self._first_text(
                    record, self.FIELD_ALIASES["amazon_sku"]
                ),
                product_name=self._first_text(
                    record, self.FIELD_ALIASES["product_name"]
                ),
                category=self._first_text(record, self.FIELD_ALIASES["category"]),
                store=self._first_text(record, self.FIELD_ALIASES["store"]),
                country=self._first_text(record, self.FIELD_ALIASES["country"]),
            )
            if not key.is_complete():
                incomplete_records += 1
                continue
            if key in result:
                duplicate_records += 1
            result[key] = result.get(key, 0) + volume
        return result, incomplete_records, duplicate_records

    def _build_updates(
        self,
        groups: dict[ProductKey, list[int]],
        period_sales: dict[int, dict[ProductKey, int]],
    ) -> tuple[
        dict[int, tuple[int, int, int, int]],
        dict[int, ProductKey],
        list[dict[str, Any]],
        int,
    ]:
        updates: dict[int, tuple[int, int, int, int]] = {}
        keys_by_row: dict[int, ProductKey] = {}
        warnings: list[dict[str, Any]] = []
        missing_rows = 0
        for key, rows in groups.items():
            found = any(key in period_sales.get(days, {}) for days in self.PERIODS)
            if not found:
                missing_rows += len(rows)
                warnings.append(
                    {
                        "level": "warning",
                        "code": "product_not_found",
                        "message": "领星结果中没有找到 A-F 完全匹配的商品，保留原值",
                        "rows": rows,
                        "identity": key.as_dict(),
                    }
                )
                continue
            sales = tuple(
                period_sales.get(days, {}).get(key, 0) for days in self.PERIODS
            )
            for row_number in rows:
                updates[row_number] = sales
                keys_by_row[row_number] = key
        return updates, keys_by_row, warnings, missing_rows

    def _write_workbook(self, plan: WorkflowPlan) -> None:
        target = plan.workbook_path
        try:
            workbook = load_workbook(target, read_only=True, data_only=False)
        except Exception as exc:
            raise LogisticsWorkflowError("无法打开本地测试表进行写入") from exc
        try:
            if self.SHEET_NAME not in workbook.sheetnames:
                raise LogisticsWorkflowError(f"工作表 {self.SHEET_NAME} 不存在")
            sheet = workbook[self.SHEET_NAME]
            for row_number, sales in plan.updates.items():
                current_key = ProductKey.from_values(
                    sheet.cell(row_number, column).value for column in range(1, 7)
                )
                if current_key != plan.keys_by_row[row_number]:
                    raise LogisticsWorkflowError(
                        f"第 {row_number} 行 A-F 已变化，已中止全部写入"
                    )
        finally:
            workbook.close()

        temp_fd, temp_name = tempfile.mkstemp(
            prefix=f".{target.stem}-",
            suffix=".xlsx",
            dir=target.parent,
        )
        os.close(temp_fd)
        try:
            self._write_cells_in_archive(target, Path(temp_name), plan.updates)
            try:
                os.replace(temp_name, target)
            except OSError as exc:
                if exc.errno != errno.EBUSY:
                    raise
                # Docker cannot replace the inode of a single-file bind mount.
                # The network workbook is disabled; this fallback only targets
                # the disposable local test copy and is followed by read-back.
                shutil.copyfile(temp_name, target)
        finally:
            if os.path.exists(temp_name):
                os.unlink(temp_name)
        self._verify_written_values(target, plan.updates)

    def _write_cells_in_archive(
        self,
        source: Path,
        destination: Path,
        updates: dict[int, tuple[int, int, int, int]],
    ) -> None:
        """Modify only target cell XML so drawings and linked content survive."""
        try:
            with zipfile.ZipFile(source, "r") as source_archive:
                sheet_path = self._sheet_xml_path(source_archive, self.SHEET_NAME)
                sheet_xml = source_archive.read(sheet_path).decode("utf-8")
                workbook_xml = self._mark_workbook_for_full_recalculation(
                    source_archive.read("xl/workbook.xml").decode("utf-8")
                )
                for row_number, sales in sorted(updates.items()):
                    for column, value in zip(self.SALES_COLUMNS, sales):
                        sheet_xml = self._replace_numeric_cell(
                            sheet_xml,
                            f"{column}{row_number}",
                            int(value),
                        )

                with zipfile.ZipFile(destination, "w") as destination_archive:
                    destination_archive.comment = source_archive.comment
                    for entry in source_archive.infolist():
                        content = source_archive.read(entry.filename)
                        if entry.filename == sheet_path:
                            content = sheet_xml.encode("utf-8")
                        elif entry.filename == "xl/workbook.xml":
                            content = workbook_xml.encode("utf-8")
                        destination_archive.writestr(entry, content)
        except LogisticsWorkflowError:
            raise
        except Exception as exc:
            raise LogisticsWorkflowError("无法安全更新本地测试表") from exc

    @staticmethod
    def _mark_workbook_for_full_recalculation(workbook_xml: str) -> str:
        """Keep formulas intact and ask Excel/WPS to recalculate the workbook on open."""
        calc_properties = '<calcPr calcId="0" calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"/>'
        pattern = re.compile(r"<calcPr\b[^>]*/>|<calcPr\b[^>]*>.*?</calcPr>", re.DOTALL)
        if pattern.search(workbook_xml):
            return pattern.sub(calc_properties, workbook_xml, count=1)
        closing_tag = "</workbook>"
        if closing_tag not in workbook_xml:
            raise LogisticsWorkflowError("工作簿结构不完整，无法设置自动重算")
        return workbook_xml.replace(closing_tag, calc_properties + closing_tag, 1)

    @staticmethod
    def _sheet_xml_path(archive: zipfile.ZipFile, sheet_name: str) -> str:
        spreadsheet_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        office_rel_ns = (
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        )
        package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
        workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        sheet_node = next(
            (
                node
                for node in workbook_root.findall(f".//{{{spreadsheet_ns}}}sheet")
                if node.attrib.get("name") == sheet_name
            ),
            None,
        )
        if sheet_node is None:
            raise LogisticsWorkflowError(f"工作表 {sheet_name} 不存在")
        relationship_id = sheet_node.attrib.get(f"{{{office_rel_ns}}}id")
        relationships_root = ElementTree.fromstring(
            archive.read("xl/_rels/workbook.xml.rels")
        )
        relationship = next(
            (
                node
                for node in relationships_root.findall(
                    f".//{{{package_rel_ns}}}Relationship"
                )
                if node.attrib.get("Id") == relationship_id
            ),
            None,
        )
        if relationship is None or not relationship.attrib.get("Target"):
            raise LogisticsWorkflowError(f"无法定位工作表 {sheet_name}")
        relationship_target = relationship.attrib["Target"]
        if relationship_target.startswith("/"):
            return relationship_target.lstrip("/")
        return posixpath.normpath(posixpath.join("xl", relationship_target))

    @classmethod
    def _replace_numeric_cell(cls, sheet_xml: str, reference: str, value: int) -> str:
        row_number_match = re.search(r"\d+$", reference)
        column_match = re.match(r"[A-Z]+", reference)
        if row_number_match is None or column_match is None:
            raise LogisticsWorkflowError(f"非法单元格坐标：{reference}")
        row_number = int(row_number_match.group())
        target_column = column_match.group()
        cell_pattern = re.compile(
            rf'<c\b(?=[^>]*\br="{re.escape(reference)}")[^>]*(?:/>|>.*?</c>)',
            re.DOTALL,
        )
        existing = cell_pattern.search(sheet_xml)
        if existing:
            replacement = cls._numeric_cell_xml(reference, value, existing.group())
            return sheet_xml[: existing.start()] + replacement + sheet_xml[existing.end() :]

        row_pattern = re.compile(
            rf'(<row\b(?=[^>]*\br="{row_number}")[^>]*>)(.*?)(</row>)',
            re.DOTALL,
        )
        row_match = row_pattern.search(sheet_xml)
        if row_match is None:
            raise LogisticsWorkflowError(f"第 {row_number} 行不存在，已中止")
        row_content = row_match.group(2)
        insertion_offset = len(row_content)
        for candidate in re.finditer(
            r'<c\b[^>]*\br="([A-Z]+)\d+"[^>]*(?:/>|>.*?</c>)',
            row_content,
            re.DOTALL,
        ):
            if cls._column_number(candidate.group(1)) > cls._column_number(target_column):
                insertion_offset = candidate.start()
                break
        new_content = (
            row_content[:insertion_offset]
            + cls._numeric_cell_xml(reference, value)
            + row_content[insertion_offset:]
        )
        return (
            sheet_xml[: row_match.start(2)]
            + new_content
            + sheet_xml[row_match.end(2) :]
        )

    @staticmethod
    def _numeric_cell_xml(reference: str, value: int, existing: str = "") -> str:
        attributes: list[str] = [f'r="{reference}"']
        opening = re.match(r"<c\b([^>]*)", existing)
        if opening:
            for name, attribute_value in re.findall(r'(\w+)="([^"]*)"', opening.group(1)):
                if name not in {"r", "t"}:
                    attributes.append(f'{name}="{attribute_value}"')
        return f'<c {" ".join(attributes)}><v>{value}</v></c>'

    @staticmethod
    def _column_number(column: str) -> int:
        number = 0
        for character in column:
            number = number * 26 + ord(character) - ord("A") + 1
        return number

    def _verify_written_values(
        self,
        target: Path,
        updates: dict[int, tuple[int, int, int, int]],
    ) -> None:
        workbook = load_workbook(target, read_only=True, data_only=False)
        try:
            sheet = workbook[self.SHEET_NAME]
            remaining = dict(updates)
            for row_number, values in enumerate(
                sheet.iter_rows(min_row=2, min_col=36, max_col=39, values_only=True),
                start=2,
            ):
                expected = remaining.pop(row_number, None)
                if expected is not None and tuple(values) != expected:
                    raise LogisticsWorkflowError(
                        f"第 {row_number} 行写入后校验失败：{tuple(values)}"
                    )
            if remaining:
                raise LogisticsWorkflowError(
                    "写入后缺少目标行：" + "、".join(map(str, remaining))
                )
        finally:
            workbook.close()

    @staticmethod
    def _validated_payload_data(payload: Any) -> Any:
        if not isinstance(payload, dict):
            raise LogisticsWorkflowError("领星销售数据格式异常")
        if payload.get("code") not in (None, 0):
            logger.warning(
                "LingXing sales payload error: code=%r message=%r data=%s",
                payload.get("code"),
                payload.get("message") or payload.get("msg"),
                repr(payload.get("data"))[:500],
            )
            raise LogisticsWorkflowError(
                payload.get("message") or payload.get("msg") or "领星查询失败"
            )
        data = payload.get("data")
        if not isinstance(data, dict):
            return data
        # LingXing wraps the actual report in data.data and uses code=1 for
        # a successful product-performance result.
        nested_code = data.get("code")
        if nested_code not in (None, 0, 1):
            logger.warning(
                "LingXing sales data error: code=%r message=%r msg=%r data=%s",
                nested_code,
                data.get("message"),
                data.get("msg"),
                repr(data)[:500],
            )
            raise LogisticsWorkflowError(data.get("msg") or data.get("message") or "领星查询失败")
        if "data" in data:
            return data.get("data")
        return data

    @staticmethod
    def _walk_dicts(value: Any) -> Iterable[dict[str, Any]]:
        if isinstance(value, dict):
            yield value
            for child in value.values():
                yield from LogisticsSalesWorkflow._walk_dicts(child)
        elif isinstance(value, list):
            for child in value:
                yield from LogisticsSalesWorkflow._walk_dicts(child)

    @staticmethod
    def _first_text(record: dict[str, Any], aliases: Iterable[str]) -> str:
        for alias in aliases:
            value = record.get(alias)
            if value not in (None, "") and not isinstance(value, (dict, list)):
                return str(value).strip()
        return ""

    @staticmethod
    def _first_number(record: dict[str, Any], aliases: Iterable[str]) -> int | None:
        for alias in aliases:
            value = record.get(alias)
            if isinstance(value, Number) and not isinstance(value, bool):
                return int(value)
            if isinstance(value, str):
                try:
                    return int(float(value.strip()))
                except ValueError:
                    continue
        return None

    @staticmethod
    def _chunks(values: list[str], size: int) -> Iterable[list[str]]:
        for start in range(0, len(values), size):
            yield values[start : start + size]

    @staticmethod
    def _file_hash(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as file:
            for chunk in iter(lambda: file.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def _platform_scope_warning() -> dict[str, Any]:
        return {
            "level": "warning",
            "code": "platform_scope_limited",
            "message": (
                "当前领星 MCP 只提供“产品表现 ASIN”销量工具；本工作流未限制国家或店铺，"
                "但该结果不能视为 Shopify、TikTok、Temu 等真正全平台销量"
            ),
            "rows": [],
        }

    def _purge_expired_previews(self, now: datetime) -> None:
        for preview_id, plan in list(self._previews.items()):
            if plan.expires_at <= now:
                self._previews.pop(preview_id, None)


logistics_sales_workflow = LogisticsSalesWorkflow()


__all__ = [
    "LogisticsSalesWorkflow",
    "LogisticsWorkflowError",
    "McpError",
    "ProductKey",
    "WorkflowPlan",
    "logistics_sales_workflow",
]
