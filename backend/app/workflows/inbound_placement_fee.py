"""Calculate Amazon US inbound placement fees for an uploaded workbook."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from math import isfinite

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


POUNDS_PER_GRAM = 1 / 453.59237
INCHES_PER_CM = 1 / 2.54
ELIGIBLE_SEGMENTS = {"小号标准尺寸", "大号标准尺寸", "小号大件", "大号大件"}


@dataclass(frozen=True)
class FeeResult:
    actual_weight_lbs: float
    volumetric_weight_lbs: float
    segment: str
    single_point_fee: float | None
    partial_split_fee: float | None


def _number(value: object) -> float | None:
    if isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if isfinite(number) and number >= 0 else None


def _band_fee(weight_lbs: float, bands: tuple[tuple[float, float, float], ...]) -> tuple[float, float]:
    for ceiling, single_point, partial_split in bands:
        if weight_lbs <= ceiling:
            return single_point, partial_split
    raise ValueError("重量超过当前尺寸分段的最高收费档")


def calculate_inbound_placement_fee(
    length_cm: object,
    width_cm: object,
    height_cm: object,
    weight_g: object,
) -> FeeResult | None:
    """Calculate the category and upper-bound fee for one sellable unit.

    Amazon's category weight is the greater of actual and volumetric weight.
    Fee ranges deliberately use the higher amount specified in the rate card.
    """

    dimensions = [_number(value) for value in (length_cm, width_cm, height_cm)]
    grams = _number(weight_g)
    if grams is None or any(value is None or value == 0 for value in dimensions):
        return None

    # The supplied Amazon workbook rounds each converted dimension to two
    # decimal places before it evaluates both the segment and volumetric weight.
    inches = sorted((round(value * INCHES_PER_CM, 2) for value in dimensions), reverse=True)
    actual_lbs = grams * POUNDS_PER_GRAM
    volumetric_lbs = inches[0] * inches[1] * inches[2] / 139
    shipping_lbs = max(actual_lbs, volumetric_lbs)
    length_plus_girth = inches[0] + 2 * (inches[1] + inches[2])

    if shipping_lbs <= 1 and inches[0] <= 15 and inches[1] <= 12 and inches[2] <= 0.75:
        segment = "小号标准尺寸"
    elif shipping_lbs <= 20 and inches[0] <= 18 and inches[1] <= 14 and inches[2] <= 8:
        segment = "大号标准尺寸"
    elif shipping_lbs <= 50 and inches[0] <= 37 and inches[1] <= 28 and inches[2] <= 20 and length_plus_girth <= 130:
        segment = "小号大件"
    elif shipping_lbs <= 50 and inches[0] <= 59 and inches[1] <= 33 and inches[2] <= 33 and length_plus_girth <= 130:
        segment = "大号大件"
    elif shipping_lbs <= 50:
        segment = "超大件（0 至 50 磅）"
    elif shipping_lbs <= 70:
        segment = "超大件（50 至 70 磅[不含 50 磅]）"
    elif shipping_lbs <= 150:
        segment = "超大件（70 至 150 磅[不含 70 磅]）"
    else:
        segment = "超大件（150 磅以上[不含 150 磅]）"

    if segment == "小号标准尺寸":
        return FeeResult(round(actual_lbs, 3), round(volumetric_lbs, 2), segment, 0.32, None)
    if segment == "大号标准尺寸":
        single_point, _ = _band_fee(
            shipping_lbs,
            ((0.75, 0.40, 0), (1.5, 0.50, 0), (3, 0.60, 0), (5, 0.76, 0), (7, 0.98, 0), (10, 1.20, 0), (15, 1.50, 0), (20, 1.90, 0)),
        )
        return FeeResult(round(actual_lbs, 3), round(volumetric_lbs, 2), segment, single_point, None)
    if segment == "小号大件":
        single_point, partial_split = _band_fee(
            shipping_lbs,
            ((5, 1.60, 1.10), (12, 2.40, 1.75), (28, 3.50, 2.19), (42, 4.95, 2.83), (50, 5.95, 3.32)),
        )
        return FeeResult(round(actual_lbs, 3), round(volumetric_lbs, 2), segment, single_point, partial_split)
    if segment == "大号大件":
        single_point, partial_split = _band_fee(
            shipping_lbs,
            ((5, 1.80, 1.25), (12, 2.90, 1.80), (28, 4.10, 2.30), (42, 5.60, 2.95), (50, 6.50, 3.50)),
        )
        return FeeResult(round(actual_lbs, 3), round(volumetric_lbs, 2), segment, single_point, partial_split)
    return FeeResult(round(actual_lbs, 3), round(volumetric_lbs, 2), segment, None, None)


def process_inbound_placement_workbook(content: bytes) -> tuple[bytes, int, int]:
    """Append fee results to the ``入库配置费`` sheet and return an XLSX file."""

    workbook = load_workbook(BytesIO(content))
    if "入库配置费" not in workbook.sheetnames:
        raise ValueError("未找到“入库配置费”工作表")
    sheet = workbook["入库配置费"]
    header_map = {str(cell.value).strip(): cell.column for cell in sheet[1] if cell.value is not None}
    required = {"单件-长(cm）", "单件-宽(cm)", "单件-高(cm)", "单件重量(g)"}
    missing = required.difference(header_map)
    if missing:
        raise ValueError(f"工作表缺少必要列：{'、'.join(sorted(missing))}")

    start_column = sheet.max_column + 1
    headers = ["计算单件重量(LBS)", "计算体积重(LBS)", "计算尺寸分段", "单点入仓费用(USD)", "部分货件拆分费用(USD)"]
    for offset, title in enumerate(headers):
        cell = sheet.cell(1, start_column + offset, title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="087B61")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[cell.column_letter].width = 20 if offset == 2 else 18

    processed = 0
    skipped = 0
    for row in range(2, sheet.max_row + 1):
        result = calculate_inbound_placement_fee(
            sheet.cell(row, header_map["单件-长(cm）"]).value,
            sheet.cell(row, header_map["单件-宽(cm)"]).value,
            sheet.cell(row, header_map["单件-高(cm)"]).value,
            sheet.cell(row, header_map["单件重量(g)"]).value,
        )
        cells = [sheet.cell(row, start_column + offset) for offset in range(5)]
        if result is None:
            for cell in cells:
                cell.value = "数据不足"
            skipped += 1
            continue
        values: list[object] = [
            result.actual_weight_lbs,
            result.volumetric_weight_lbs,
            result.segment,
            result.single_point_fee if result.segment in ELIGIBLE_SEGMENTS else "不适用（超大件）",
            result.partial_split_fee if result.partial_split_fee is not None else "—",
        ]
        for cell, value in zip(cells, values):
            cell.value = value
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(value, float):
                cell.number_format = '0.00'
        processed += 1

    sheet.freeze_panes = "A2"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue(), processed, skipped
