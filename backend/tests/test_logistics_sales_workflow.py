import tempfile
import unittest
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook

from app.workflows.logistics_sales_workflow import (
    LogisticsSalesWorkflow,
    LogisticsWorkflowError,
    McpError,
    ProductKey,
    WorkflowPlan,
)
from app.workflows.cached_logistics_sales_workflow import CachedLogisticsSalesWorkflow
from app.workflows.sales_cache import DailySalesRecord, SalesCache


class LogisticsSalesWorkflowTests(unittest.TestCase):
    def setUp(self):
        self.workflow = LogisticsSalesWorkflow()
        self.key = ProductKey(
            sku="60003-2",
            amazon_sku="60003US05HXD",
            product_name='三头肌绳 Tricep Rope 36" 灰色（Gray）',
            category="三头肌绳",
            store="HXD-ERGO",
            country="美国",
        )

    def _create_workbook(self, target: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        headers = [
            "领星SKU",
            "亚马逊SKU",
            "品名",
            "品类",
            "店铺",
            "国家",
        ]
        for column, header in enumerate(headers, start=1):
            sheet.cell(1, column).value = header
        for row_number in (2, 5):
            for column, value in enumerate(self.key.as_dict().values(), start=1):
                sheet.cell(row_number, column).value = value
        sheet["AR2"] = "=IF(AJ2>0,AJ2,0)"
        missing = ProductKey(
            sku="NO-MATCH",
            amazon_sku="NO-MATCH-US",
            product_name="无匹配商品",
            category="测试",
            store="TEST",
            country="美国",
        )
        for column, value in enumerate(missing.as_dict().values(), start=1):
            sheet.cell(3, column).value = value
        sheet["AJ3"] = 88
        workbook.save(target)
        workbook.close()

    def test_extracts_and_aggregates_exact_a_to_f_records(self):
        payload = {
            "code": 0,
            "data": {
                "list": [
                    {
                        "local_sku": self.key.sku,
                        "msku": self.key.amazon_sku,
                        "local_name": self.key.product_name,
                        "category_name": self.key.category,
                        "seller_name": self.key.store,
                        "country_name": self.key.country,
                        "volume": 6,
                    },
                    {
                        "local_sku": self.key.sku,
                        "msku": self.key.amazon_sku,
                        "local_name": self.key.product_name,
                        "category_name": self.key.category,
                        "seller_name": self.key.store,
                        "country_name": self.key.country,
                        "volume": 4,
                    },
                ]
            },
        }
        result, incomplete, duplicate = self.workflow._extract_sales_map(payload)
        self.assertEqual(result[self.key], 10)
        self.assertEqual(incomplete, 0)
        self.assertEqual(duplicate, 1)

    def test_rejects_lingxing_permission_error(self):
        payload = {
            "code": 0,
            "data": {"code": 8002, "msg": "暂无产品表现查看权限"},
        }
        with self.assertRaisesRegex(LogisticsWorkflowError, "暂无产品表现查看权限"):
            self.workflow._extract_sales_map(payload)

    def test_empty_lingxing_error_message_has_readable_fallback(self):
        with self.assertRaisesRegex(LogisticsWorkflowError, "领星查询失败"):
            self.workflow._extract_sales_map(
                {"code": 0, "data": {"code": 8002, "msg": None}}
            )

    def test_empty_mcp_error_message_has_readable_fallback(self):
        self.assertEqual(str(McpError(None)), "领星 MCP 调用失败")

    def test_unwraps_lingxing_nested_code_one_success_payload(self):
        payload = {
            "code": 0,
            "data": {
                "code": 1,
                "msg": None,
                "trace_id": "trace-test",
                "data": {"list": [{"local_sku": "60003-2", "volume": 6}]},
            },
        }
        self.assertEqual(
            self.workflow._validated_payload_data(payload),
            {"list": [{"local_sku": "60003-2", "volume": 6}]},
        )

    def test_duplicate_rows_receive_same_values_and_missing_row_is_untouched(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            target = Path(temp_dir) / "备货逻辑看板表-工作流测试.xlsx"
            self._create_workbook(target)
            groups, warnings, total_rows = self.workflow._read_product_groups(target)
            self.assertEqual(total_rows, 3)
            self.assertEqual(groups[self.key], [2, 5])
            self.assertTrue(any(item["code"] == "duplicate_identity" for item in warnings))

            period_sales = {
                3: {self.key: 10},
                7: {self.key: 20},
                15: {self.key: 30},
                30: {self.key: 40},
            }
            updates, keys_by_row, match_warnings, missing_rows = (
                self.workflow._build_updates(groups, period_sales)
            )
            self.assertEqual(updates[2], (10, 20, 30, 40))
            self.assertEqual(updates[5], (10, 20, 30, 40))
            self.assertEqual(missing_rows, 1)
            self.assertTrue(
                any(item["code"] == "product_not_found" for item in match_warnings)
            )

            now = datetime.now(ZoneInfo("Asia/Shanghai"))
            plan = WorkflowPlan(
                preview_id="test",
                workbook_path=target,
                workbook_hash=self.workflow._file_hash(target),
                expires_at=now + timedelta(minutes=15),
                updates=updates,
                keys_by_row=keys_by_row,
                total_rows=total_rows,
                unique_products=len(groups),
                missing_rows=missing_rows,
                duplicate_groups=1,
                warnings=warnings + match_warnings,
            )
            with zipfile.ZipFile(target, "r") as archive:
                sheet_path = self.workflow._sheet_xml_path(archive, "Sheet1")
                original_parts = {
                    entry.filename: archive.read(entry.filename)
                    for entry in archive.infolist()
                }
            self.workflow._write_workbook(plan)

            with zipfile.ZipFile(target, "r") as archive:
                written_parts = {
                    entry.filename: archive.read(entry.filename)
                    for entry in archive.infolist()
                }
            self.assertEqual(original_parts.keys(), written_parts.keys())
            for part_name, content in original_parts.items():
                if part_name == "xl/workbook.xml":
                    self.assertNotEqual(content, written_parts[part_name])
                    self.assertIn(b'calcMode="auto"', written_parts[part_name])
                    self.assertIn(b'fullCalcOnLoad="1"', written_parts[part_name])
                    self.assertIn(b'forceFullCalc="1"', written_parts[part_name])
                elif part_name != sheet_path:
                    self.assertEqual(content, written_parts[part_name])

            result = load_workbook(target, data_only=False)
            try:
                sheet = result["Sheet1"]
                for row_number in (2, 5):
                    self.assertEqual(
                        tuple(
                            sheet[f"{column}{row_number}"].value
                            for column in self.workflow.SALES_COLUMNS
                        ),
                        (10, 20, 30, 40),
                    )
                self.assertEqual(sheet["AR2"].value, "=IF(AJ2>0,AJ2,0)")
                self.assertEqual(sheet["AJ3"].value, 88)
            finally:
                result.close()

    def test_full_recalculation_marker_preserves_formula_text(self):
        workbook_xml = '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><calcPr calcId="191029"/></workbook>'
        marked = self.workflow._mark_workbook_for_full_recalculation(workbook_xml)
        self.assertIn('calcMode="auto"', marked)
        self.assertIn('fullCalcOnLoad="1"', marked)
        self.assertIn('forceFullCalc="1"', marked)


    def test_platform_scope_warning_is_explicit(self):
        warning = self.workflow._platform_scope_warning()
        self.assertEqual(warning["code"], "platform_scope_limited")
        self.assertIn("真正全平台", warning["message"])


class SalesCacheTests(unittest.TestCase):
    def test_coverage_tracks_missing_dates_and_daily_rows(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            cache = SalesCache(Path(temp_dir) / "sales.sqlite3")
            start = datetime(2026, 8, 20).date()
            end = datetime(2026, 8, 22).date()
            self.assertEqual(
                cache.missing_dates(["SKU-1"], start, end)["SKU-1"],
                [start, start + timedelta(days=1), end],
            )
            cache.save_daily_records(
                [
                    DailySalesRecord(
                        sales_date="2026-08-20",
                        sku="SKU-1",
                        amazon_sku="",
                        product_name="商品",
                        category="品类",
                        store="店铺",
                        country="美国",
                        platform="其他",
                        volume=7,
                        trace_id="trace-1",
                    )
                ],
                ["SKU-1"],
                start,
                end,
                "trace-1",
            )
            self.assertEqual(cache.missing_dates(["SKU-1"], start, end), {})
            self.assertEqual(
                cache.daily_records("SKU-1", start, end)[0].volume,
                7,
            )

    def test_cached_matching_allows_empty_amazon_sku(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workflow = CachedLogisticsSalesWorkflow()
            workflow.cache = SalesCache(Path(temp_dir) / "sales.sqlite3")
            key = ProductKey(
                sku="SKU-1",
                amazon_sku="",
                product_name="商品",
                category="品类",
                store="店铺",
                country="美国",
            )
            workflow.cache.save_daily_records(
                [
                    DailySalesRecord(
                        sales_date="2026-08-24",
                        sku="SKU-1",
                        amazon_sku="",
                        product_name="商品",
                        category="品类",
                        store="店铺",
                        country="美国",
                        platform="其他",
                        volume=9,
                        trace_id="trace-2",
                    )
                ],
                ["SKU-1"],
                datetime(2026, 8, 24).date(),
                datetime(2026, 8, 24).date(),
                "trace-2",
            )
            updates, _, warnings, missing = workflow._build_cached_updates(
                {key: [2]}, datetime(2026, 8, 25).date()
            )
            self.assertEqual(missing, 0)
            self.assertEqual(updates[2], (9, 9, 9, 9))
            self.assertFalse(warnings)


if __name__ == "__main__":
    unittest.main()
